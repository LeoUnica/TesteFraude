import uuid
import json
import io
import csv
from datetime import datetime
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ...database import get_db
from ...models.proposal import Proposal
from ...core.audit import audit_log
from ...api.deps import get_current_user
from ...models.user import User
import structlog

logger = structlog.get_logger()

router = APIRouter(prefix="/proposals", tags=["proposals"])


class ProposalCreate(BaseModel):
    code: str
    cpf: str
    client_name: str
    broker_id: Optional[str] = None
    convenio_id: Optional[str] = None
    bank_id: Optional[str] = None
    product_id: Optional[str] = None
    product_fgts: Optional[str] = None
    value: float = 0.0
    installments: int = 0
    status: str = "Pendente"
    antifraud_status: str = "Nao Analisado"
    pipeline_status: Optional[str] = None
    pipeline_phase: Optional[str] = None
    import_date: Optional[datetime] = None
    proposal_date: Optional[datetime] = None
    notes: Optional[str] = None
    documents: list = []
    history: list = []


class ProposalUpdate(BaseModel):
    cpf: Optional[str] = None
    client_name: Optional[str] = None
    broker_id: Optional[str] = None
    convenio_id: Optional[str] = None
    bank_id: Optional[str] = None
    product_id: Optional[str] = None
    product_fgts: Optional[str] = None
    value: Optional[float] = None
    installments: Optional[int] = None
    status: Optional[str] = None
    antifraud_status: Optional[str] = None
    pipeline_status: Optional[str] = None
    pipeline_phase: Optional[str] = None
    import_date: Optional[datetime] = None
    proposal_date: Optional[datetime] = None
    notes: Optional[str] = None
    documents: Optional[list] = None
    history: Optional[list] = None


class EndorseRequest(BaseModel):
    endorsement_date: Optional[datetime] = None


class ChangeConvenioRequest(BaseModel):
    convenio_id: str


class ChangeBrokerRequest(BaseModel):
    broker_id: str


def proposal_to_dict(p: Proposal) -> dict:
    try:
        docs = json.loads(p.documents or "[]")
    except Exception:
        docs = []
    try:
        hist = json.loads(p.history or "[]")
    except Exception:
        hist = []
    return {
        "id": p.id,
        "code": p.code,
        "cpf": p.cpf,
        "client_name": p.client_name,
        "broker_id": p.broker_id,
        "convenio_id": p.convenio_id,
        "bank_id": p.bank_id,
        "product_id": p.product_id,
        "product_fgts": p.product_fgts,
        "value": p.value,
        "installments": p.installments,
        "status": p.status,
        "antifraud_status": p.antifraud_status,
        "pipeline_status": p.pipeline_status,
        "pipeline_phase": p.pipeline_phase,
        "import_date": p.import_date.isoformat() if p.import_date else None,
        "endorsement_date": p.endorsement_date.isoformat() if p.endorsement_date else None,
        "proposal_date": p.proposal_date.isoformat() if p.proposal_date else None,
        "notes": p.notes,
        "documents": docs,
        "history": hist,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        "created_by": p.created_by,
        "imported_by": p.imported_by,
    }


@router.get("/")
async def list_proposals(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    status_filter: Optional[str] = Query(None, alias="status"),
    antifraud_status: Optional[str] = None,
    broker_id: Optional[str] = None,
    convenio_id: Optional[str] = None,
    bank_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
):
    query = db.query(Proposal)
    if status_filter:
        query = query.filter(Proposal.status == status_filter)
    if antifraud_status:
        query = query.filter(Proposal.antifraud_status == antifraud_status)
    if broker_id:
        query = query.filter(Proposal.broker_id == broker_id)
    if convenio_id:
        query = query.filter(Proposal.convenio_id == convenio_id)
    if bank_id:
        query = query.filter(Proposal.bank_id == bank_id)
    if date_from:
        try:
            df = datetime.fromisoformat(date_from)
            query = query.filter(Proposal.created_at >= df)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.fromisoformat(date_to)
            query = query.filter(Proposal.created_at <= dt)
        except ValueError:
            pass
    if search:
        query = query.filter(
            (Proposal.client_name.ilike(f"%{search}%"))
            | (Proposal.cpf.ilike(f"%{search}%"))
            | (Proposal.code.ilike(f"%{search}%"))
        )

    total = query.count()
    proposals = (
        query.order_by(Proposal.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return {
        "data": [proposal_to_dict(p) for p in proposals],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_proposal(
    body: ProposalCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    existing = db.query(Proposal).filter(Proposal.code == body.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Código de proposta já cadastrado")

    proposal = Proposal(
        id=str(uuid.uuid4()),
        code=body.code,
        cpf=body.cpf,
        client_name=body.client_name,
        broker_id=body.broker_id,
        convenio_id=body.convenio_id,
        bank_id=body.bank_id,
        product_id=body.product_id,
        product_fgts=body.product_fgts,
        value=body.value,
        installments=body.installments,
        status=body.status,
        antifraud_status=body.antifraud_status,
        pipeline_status=body.pipeline_status,
        pipeline_phase=body.pipeline_phase,
        import_date=body.import_date,
        proposal_date=body.proposal_date,
        notes=body.notes,
        documents=json.dumps(body.documents),
        history=json.dumps(body.history),
        created_by=current_user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(proposal)
    db.commit()
    db.refresh(proposal)

    audit_log(
        db, current_user.id, current_user.name,
        "CREATE_PROPOSAL", "proposals", entity_id=proposal.id, entity_type="Proposal",
        new_value=json.dumps({"code": proposal.code, "client_name": proposal.client_name}),
    )

    return proposal_to_dict(proposal)


@router.put("/{proposal_id}")
async def update_proposal(
    proposal_id: str,
    body: ProposalUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    proposal = db.query(Proposal).filter(Proposal.id == proposal_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")

    old_data = proposal_to_dict(proposal)

    update_data = body.model_dump(exclude_unset=True)
    if "documents" in update_data:
        update_data["documents"] = json.dumps(update_data["documents"])
    if "history" in update_data:
        update_data["history"] = json.dumps(update_data["history"])

    for field, value in update_data.items():
        setattr(proposal, field, value)
    proposal.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(proposal)

    audit_log(
        db, current_user.id, current_user.name,
        "UPDATE_PROPOSAL", "proposals", entity_id=proposal.id, entity_type="Proposal",
        old_value=json.dumps(old_data), new_value=json.dumps(proposal_to_dict(proposal)),
    )

    return proposal_to_dict(proposal)


@router.patch("/{proposal_id}/endorse")
async def endorse_proposal(
    proposal_id: str,
    body: EndorseRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    proposal = db.query(Proposal).filter(Proposal.id == proposal_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")

    proposal.endorsement_date = body.endorsement_date or datetime.utcnow()
    proposal.updated_at = datetime.utcnow()
    db.commit()

    audit_log(
        db, current_user.id, current_user.name,
        "ENDORSE_PROPOSAL", "proposals", entity_id=proposal.id, entity_type="Proposal",
        new_value=json.dumps({"endorsement_date": proposal.endorsement_date.isoformat()}),
    )

    return proposal_to_dict(proposal)


@router.patch("/{proposal_id}/change-convenio")
async def change_proposal_convenio(
    proposal_id: str,
    body: ChangeConvenioRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    proposal = db.query(Proposal).filter(Proposal.id == proposal_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")

    old_convenio = proposal.convenio_id
    proposal.convenio_id = body.convenio_id
    proposal.updated_at = datetime.utcnow()
    db.commit()

    audit_log(
        db, current_user.id, current_user.name,
        "CHANGE_CONVENIO", "proposals", entity_id=proposal.id, entity_type="Proposal",
        old_value=json.dumps({"convenio_id": old_convenio}),
        new_value=json.dumps({"convenio_id": body.convenio_id}),
    )

    return proposal_to_dict(proposal)


@router.post("/import")
async def import_proposals(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    content = await file.read()
    rows = []

    if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        decoded = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = [{k.strip().lower(): v.strip() for k, v in r.items()} for r in reader]

    inserted, skipped, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        code = row.get("codigo") or row.get("code") or row.get("proposta") or row.get("numero_proposta") or ""
        cpf = row.get("cpf") or row.get("documento") or ""
        client_name = row.get("nome") or row.get("cliente") or row.get("client_name") or ""

        if not code or not cpf or not client_name:
            errors.append(f"Linha {i}: código, CPF e nome são obrigatórios")
            skipped += 1
            continue

        existing = db.query(Proposal).filter(Proposal.code == code).first()
        if existing:
            skipped += 1
            continue

        try:
            value = float(row.get("valor") or row.get("value") or 0)
        except (ValueError, TypeError):
            value = 0.0

        db.add(Proposal(
            id=str(uuid.uuid4()),
            code=code,
            cpf=cpf,
            client_name=client_name,
            value=value,
            status=row.get("status") or "Pendente",
            antifraud_status="Nao Analisado",
            import_date=datetime.utcnow(),
            documents=json.dumps([]),
            history=json.dumps([]),
            imported_by=current_user.id,
            created_by=current_user.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        ))
        inserted += 1

    db.commit()
    audit_log(db, current_user.id, current_user.name, "IMPORT_PROPOSALS", "proposals",
              new_value=json.dumps({"inserted": inserted, "skipped": skipped}))

    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@router.get("/export/excel")
async def export_proposals_excel(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    status_filter: Optional[str] = Query(None, alias="status"),
    antifraud_status: Optional[str] = None,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = db.query(Proposal)
    if status_filter:
        query = query.filter(Proposal.status == status_filter)
    if antifraud_status:
        query = query.filter(Proposal.antifraud_status == antifraud_status)
    proposals = query.order_by(Proposal.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propostas"

    headers = ["Código", "CPF", "Cliente", "Valor", "Status", "Status Antifraude", "Data Importação", "Data Criação"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for p in proposals:
        ws.append([
            p.code, p.cpf, p.client_name, p.value, p.status, p.antifraud_status,
            p.import_date.strftime("%d/%m/%Y") if p.import_date else "",
            p.created_at.strftime("%d/%m/%Y") if p.created_at else "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=propostas.xlsx"},
    )


@router.patch("/{proposal_id}/change-broker")
async def change_proposal_broker(
    proposal_id: str,
    body: ChangeBrokerRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    proposal = db.query(Proposal).filter(Proposal.id == proposal_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")

    old_broker = proposal.broker_id
    proposal.broker_id = body.broker_id
    proposal.updated_at = datetime.utcnow()
    db.commit()

    audit_log(
        db, current_user.id, current_user.name,
        "CHANGE_BROKER", "proposals", entity_id=proposal.id, entity_type="Proposal",
        old_value=json.dumps({"broker_id": old_broker}),
        new_value=json.dumps({"broker_id": body.broker_id}),
    )

    return proposal_to_dict(proposal)
