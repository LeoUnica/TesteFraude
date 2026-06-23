import uuid
import json
from datetime import datetime
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ...database import get_db
from ...models.blacklist import BlacklistEntry
from ...core.audit import audit_log
from ...api.deps import get_current_user
from ...models.user import User
import structlog
import io

logger = structlog.get_logger()

router = APIRouter(prefix="/blacklist", tags=["blacklist"])


class BlacklistCreate(BaseModel):
    type: str  # cpf, cnpj, phone, email
    value: str
    reason: Optional[str] = None
    source: Optional[str] = None


class BlacklistUpdate(BaseModel):
    reason: Optional[str] = None
    source: Optional[str] = None
    status: Optional[str] = None


def entry_to_dict(e: BlacklistEntry) -> dict:
    return {
        "id": e.id,
        "type": e.type,
        "value": e.value,
        "reason": e.reason,
        "source": e.source,
        "status": e.status,
        "created_by": e.created_by,
        "created_at": e.created_at.isoformat() if e.created_at else None,
        "updated_at": e.updated_at.isoformat() if e.updated_at else None,
    }


@router.get("/")
async def list_blacklist(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=200),
    type_filter: Optional[str] = Query(None, alias="type"),
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
):
    query = db.query(BlacklistEntry)
    if type_filter:
        query = query.filter(BlacklistEntry.type == type_filter)
    if status_filter:
        query = query.filter(BlacklistEntry.status == status_filter)
    if search:
        query = query.filter(BlacklistEntry.value.ilike(f"%{search}%"))

    total = query.count()
    entries = query.order_by(BlacklistEntry.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "data": [entry_to_dict(e) for e in entries],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.post("/check")
async def check_blacklist(
    body: BlacklistCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    entry = db.query(BlacklistEntry).filter(
        BlacklistEntry.value == body.value,
        BlacklistEntry.status == "ativo",
    ).first()
    return {"blocked": entry is not None, "entry": entry_to_dict(entry) if entry else None}


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_entry(
    body: BlacklistCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    existing = db.query(BlacklistEntry).filter(BlacklistEntry.value == body.value).first()
    if existing:
        raise HTTPException(status_code=400, detail="Valor já consta na blacklist")

    entry = BlacklistEntry(
        id=str(uuid.uuid4()),
        type=body.type,
        value=body.value,
        reason=body.reason,
        source=body.source,
        created_by=current_user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)

    audit_log(db, current_user.id, current_user.name, "CREATE_BLACKLIST", "blacklist",
              entity_id=entry.id, new_value=json.dumps({"type": entry.type, "value": entry.value}))

    return entry_to_dict(entry)


@router.put("/{entry_id}")
async def update_entry(
    entry_id: str,
    body: BlacklistUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    entry = db.query(BlacklistEntry).filter(BlacklistEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entrada não encontrada")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(entry, field, value)
    entry.updated_at = datetime.utcnow()
    db.commit()

    audit_log(db, current_user.id, current_user.name, "UPDATE_BLACKLIST", "blacklist", entity_id=entry.id)
    return entry_to_dict(entry)


@router.delete("/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_entry(
    entry_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    entry = db.query(BlacklistEntry).filter(BlacklistEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entrada não encontrada")
    db.delete(entry)
    db.commit()
    audit_log(db, current_user.id, current_user.name, "DELETE_BLACKLIST", "blacklist", entity_id=entry_id)


@router.post("/import")
async def import_blacklist(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    import csv

    content = await file.read()
    inserted = 0
    skipped = 0
    errors = []

    rows = []
    if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        decoded = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = [{k.strip().lower(): v.strip() for k, v in r.items()} for r in reader]

    for i, row in enumerate(rows, start=2):
        value = row.get("valor") or row.get("value") or row.get("cpf") or row.get("cnpj") or row.get("telefone") or ""
        tipo = row.get("tipo") or row.get("type") or "cpf"
        reason = row.get("motivo") or row.get("reason") or ""

        if not value:
            errors.append(f"Linha {i}: valor vazio")
            skipped += 1
            continue

        existing = db.query(BlacklistEntry).filter(BlacklistEntry.value == value).first()
        if existing:
            skipped += 1
            continue

        db.add(BlacklistEntry(
            id=str(uuid.uuid4()),
            type=tipo,
            value=value,
            reason=reason,
            created_by=current_user.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        ))
        inserted += 1

    db.commit()
    audit_log(db, current_user.id, current_user.name, "IMPORT_BLACKLIST", "blacklist",
              new_value=json.dumps({"inserted": inserted, "skipped": skipped}))

    return {"inserted": inserted, "skipped": skipped, "errors": errors}
