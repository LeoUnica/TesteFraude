import uuid
import json
import io
import csv
from datetime import datetime
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ...database import get_db
from ...models.broker import Broker
from ...core.audit import audit_log
from ...api.deps import get_current_user
from ...models.user import User
import structlog

logger = structlog.get_logger()

router = APIRouter(prefix="/brokers", tags=["brokers"])


class BrokerCreate(BaseModel):
    code: str
    name: str
    cpf_cnpj: str
    type: str = "Externo"
    status: str = "ativo"
    group_id: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None


class BrokerUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    cpf_cnpj: Optional[str] = None
    type: Optional[str] = None
    status: Optional[str] = None
    group_id: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None


def broker_to_dict(b: Broker) -> dict:
    return {
        "id": b.id,
        "code": b.code,
        "name": b.name,
        "cpf_cnpj": b.cpf_cnpj,
        "type": b.type,
        "status": b.status,
        "group_id": b.group_id,
        "email": b.email,
        "phone": b.phone,
        "address": b.address,
        "city": b.city,
        "state": b.state,
        "created_at": b.created_at.isoformat() if b.created_at else None,
        "updated_at": b.updated_at.isoformat() if b.updated_at else None,
        "created_by": b.created_by,
    }


@router.get("/")
async def list_brokers(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    status_filter: Optional[str] = Query(None, alias="status"),
    group_id: Optional[str] = None,
    search: Optional[str] = None,
):
    query = db.query(Broker)
    if status_filter:
        query = query.filter(Broker.status == status_filter)
    if group_id:
        query = query.filter(Broker.group_id == group_id)
    if search:
        query = query.filter(
            (Broker.name.ilike(f"%{search}%"))
            | (Broker.code.ilike(f"%{search}%"))
            | (Broker.cpf_cnpj.ilike(f"%{search}%"))
        )

    total = query.count()
    brokers = query.order_by(Broker.name).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "data": [broker_to_dict(b) for b in brokers],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_broker(
    body: BrokerCreate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    existing = db.query(Broker).filter(
        (Broker.code == body.code) | (Broker.cpf_cnpj == body.cpf_cnpj)
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Código ou CPF/CNPJ já cadastrado")

    broker = Broker(
        id=str(uuid.uuid4()),
        code=body.code,
        name=body.name,
        cpf_cnpj=body.cpf_cnpj,
        type=body.type,
        status=body.status,
        group_id=body.group_id,
        email=body.email,
        phone=body.phone,
        address=body.address,
        city=body.city,
        state=body.state,
        created_by=current_user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(broker)
    db.commit()
    db.refresh(broker)

    audit_log(
        db, current_user.id, current_user.name,
        "CREATE_BROKER", "brokers", entity_id=broker.id, entity_type="Broker",
        new_value=json.dumps({"name": broker.name, "code": broker.code}),
    )

    return broker_to_dict(broker)


@router.put("/{broker_id}")
async def update_broker(
    broker_id: str,
    body: BrokerUpdate,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    broker = db.query(Broker).filter(Broker.id == broker_id).first()
    if not broker:
        raise HTTPException(status_code=404, detail="Correspondente não encontrado")

    old_data = broker_to_dict(broker)

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(broker, field, value)
    broker.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(broker)

    audit_log(
        db, current_user.id, current_user.name,
        "UPDATE_BROKER", "brokers", entity_id=broker.id, entity_type="Broker",
        old_value=json.dumps(old_data), new_value=json.dumps(broker_to_dict(broker)),
    )

    return broker_to_dict(broker)


@router.patch("/{broker_id}/toggle-status")
async def toggle_broker_status(
    broker_id: str,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    broker = db.query(Broker).filter(Broker.id == broker_id).first()
    if not broker:
        raise HTTPException(status_code=404, detail="Correspondente não encontrado")

    broker.status = "inativo" if broker.status == "ativo" else "ativo"
    broker.updated_at = datetime.utcnow()
    db.commit()

    audit_log(
        db, current_user.id, current_user.name,
        "TOGGLE_BROKER_STATUS", "brokers", entity_id=broker.id, entity_type="Broker",
        new_value=json.dumps({"status": broker.status}),
    )

    return broker_to_dict(broker)


@router.post("/import")
async def import_brokers(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    content = await file.read()
    rows = []

    if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        decoded = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = [{k.strip().lower(): v.strip() for k, v in r.items()} for r in reader]

    inserted, skipped, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        code = row.get("codigo") or row.get("code") or row.get("cód") or ""
        name = row.get("nome") or row.get("name") or ""
        cpf_cnpj = row.get("cpf_cnpj") or row.get("cpf/cnpj") or row.get("documento") or ""

        if not code or not name or not cpf_cnpj:
            errors.append(f"Linha {i}: campos obrigatórios ausentes (código, nome, cpf_cnpj)")
            skipped += 1
            continue

        existing = db.query(Broker).filter(
            (Broker.code == code) | (Broker.cpf_cnpj == cpf_cnpj)
        ).first()
        if existing:
            skipped += 1
            continue

        db.add(Broker(
            id=str(uuid.uuid4()),
            code=code,
            name=name,
            cpf_cnpj=cpf_cnpj,
            type=row.get("tipo") or row.get("type") or "Externo",
            email=row.get("email") or None,
            phone=row.get("telefone") or row.get("phone") or None,
            city=row.get("cidade") or row.get("city") or None,
            state=row.get("estado") or row.get("state") or None,
            status="ativo",
            created_by=current_user.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        ))
        inserted += 1

    db.commit()
    audit_log(db, current_user.id, current_user.name, "IMPORT_BROKERS", "brokers",
              new_value=json.dumps({"inserted": inserted, "skipped": skipped}))

    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@router.get("/export/excel")
async def export_brokers_excel(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    status_filter: Optional[str] = Query(None, alias="status"),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    brokers = db.query(Broker)
    if status_filter:
        brokers = brokers.filter(Broker.status == status_filter)
    brokers = brokers.order_by(Broker.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corretores"

    headers = ["Código", "Nome", "CPF/CNPJ", "Tipo", "Status", "E-mail", "Telefone", "Cidade", "Estado"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, b in enumerate(brokers, 2):
        ws.append([b.code, b.name, b.cpf_cnpj, b.type, b.status, b.email or "", b.phone or "", b.city or "", b.state or ""])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=corretores.xlsx"},
    )
