import json
import io
from datetime import datetime
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ...database import get_db
from ...models.proposal import Proposal
from ...models.broker import Broker
from ...models.antifraud import AntifraudAnalysis
from ...models.audit import AuditLog
from ...api.deps import get_current_user
from ...models.user import User

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/proposals")
async def proposals_report(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    status_filter: Optional[str] = Query(None, alias="status"),
    antifraud_status: Optional[str] = None,
    broker_id: Optional[str] = None,
    bank_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    query = db.query(Proposal)
    if status_filter:
        query = query.filter(Proposal.status == status_filter)
    if antifraud_status:
        query = query.filter(Proposal.antifraud_status == antifraud_status)
    if broker_id:
        query = query.filter(Proposal.broker_id == broker_id)
    if bank_id:
        query = query.filter(Proposal.bank_id == bank_id)
    if date_from:
        try:
            query = query.filter(Proposal.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Proposal.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass

    total = query.count()
    proposals = (
        query.order_by(Proposal.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    data = []
    for p in proposals:
        data.append({
            "id": p.id,
            "code": p.code,
            "cpf": p.cpf,
            "client_name": p.client_name,
            "broker_id": p.broker_id,
            "bank_id": p.bank_id,
            "convenio_id": p.convenio_id,
            "product_id": p.product_id,
            "value": p.value,
            "installments": p.installments,
            "status": p.status,
            "antifraud_status": p.antifraud_status,
            "endorsement_date": p.endorsement_date.isoformat() if p.endorsement_date else None,
            "import_date": p.import_date.isoformat() if p.import_date else None,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        })

    return {"data": data, "total": total, "page": page, "per_page": per_page}


@router.get("/proposals/excel")
async def proposals_excel(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    status_filter: Optional[str] = Query(None, alias="status"),
    antifraud_status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = db.query(Proposal)
    if status_filter:
        query = query.filter(Proposal.status == status_filter)
    if antifraud_status:
        query = query.filter(Proposal.antifraud_status == antifraud_status)
    if date_from:
        try:
            query = query.filter(Proposal.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Proposal.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass
    proposals = query.order_by(Proposal.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propostas"
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    bold_white = Font(color="FFFFFF", bold=True)

    headers = ["Código", "CPF", "Cliente", "Valor (R$)", "Status", "Antifraude", "Importado em", "Criado em"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center")

    for p in proposals:
        ws.append([
            p.code, p.cpf, p.client_name, p.value, p.status, p.antifraud_status,
            p.import_date.strftime("%d/%m/%Y") if p.import_date else "",
            p.created_at.strftime("%d/%m/%Y %H:%M") if p.created_at else "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_propostas.xlsx"})


@router.get("/proposals/pdf")
async def proposals_pdf(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    status_filter: Optional[str] = Query(None, alias="status"),
    antifraud_status: Optional[str] = None,
):
    from fpdf import FPDF

    query = db.query(Proposal)
    if status_filter:
        query = query.filter(Proposal.status == status_filter)
    if antifraud_status:
        query = query.filter(Proposal.antifraud_status == antifraud_status)
    proposals = query.order_by(Proposal.created_at.desc()).limit(500).all()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Relatório de Propostas", ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(proposals)}", ln=True, align="C")
    pdf.ln(3)

    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    cols = [("Código", 35), ("CPF", 32), ("Cliente", 70), ("Valor", 25), ("Status", 30), ("Antifraude", 35), ("Data", 25)]
    for label, w in cols:
        pdf.cell(w, 7, label, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 7)
    for i, p in enumerate(proposals):
        fill_row = i % 2 == 0
        pdf.set_fill_color(240, 244, 248) if fill_row else pdf.set_fill_color(255, 255, 255)
        row = [
            (p.code or "", 35), (p.cpf or "", 32), (p.client_name or "", 70),
            (f"R$ {p.value:,.2f}" if p.value else "R$ 0,00", 25),
            (p.status or "", 30), (p.antifraud_status or "", 35),
            (p.created_at.strftime("%d/%m/%Y") if p.created_at else "", 25),
        ]
        for val, w in row:
            pdf.cell(w, 6, str(val)[:30], border=1, fill=True)
        pdf.ln()

    out = io.BytesIO(pdf.output())
    return StreamingResponse(out, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_propostas.pdf"})


@router.get("/brokers/excel")
async def brokers_excel(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    brokers = db.query(Broker).order_by(Broker.name).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corretores"
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    headers = ["Código", "Nome", "CPF/CNPJ", "Tipo", "Status", "E-mail", "Telefone", "Cidade", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for b in brokers:
        ws.append([b.code, b.name, b.cpf_cnpj, b.type, b.status, b.email or "", b.phone or "", b.city or "", b.state or ""])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_corretores.xlsx"})


@router.get("/brokers")
async def brokers_report(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    status_filter: Optional[str] = Query(None, alias="status"),
):
    query = db.query(Broker)
    if status_filter:
        query = query.filter(Broker.status == status_filter)

    total = query.count()
    brokers = query.order_by(Broker.name).offset((page - 1) * per_page).limit(per_page).all()

    data = [
        {
            "id": b.id,
            "code": b.code,
            "name": b.name,
            "cpf_cnpj": b.cpf_cnpj,
            "type": b.type,
            "status": b.status,
            "group_id": b.group_id,
            "email": b.email,
            "city": b.city,
            "state": b.state,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in brokers
    ]

    return {"data": data, "total": total, "page": page, "per_page": per_page}


@router.get("/antifraud")
async def antifraud_report(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
):
    query = db.query(AntifraudAnalysis).order_by(AntifraudAnalysis.created_at.desc())
    total = query.count()
    analyses = query.offset((page - 1) * per_page).limit(per_page).all()

    data = []
    for a in analyses:
        proposal = db.query(Proposal).filter(Proposal.id == a.proposal_id).first()
        data.append({
            "id": a.id,
            "proposal_id": a.proposal_id,
            "proposal_code": proposal.code if proposal else None,
            "client_name": proposal.client_name if proposal else None,
            "rule_id": a.rule_id,
            "analyst_id": a.analyst_id,
            "status": a.status,
            "notes": a.notes,
            "schedule_date": a.schedule_date.isoformat() if a.schedule_date else None,
            "created_at": a.created_at.isoformat() if a.created_at else None,
        })

    return {"data": data, "total": total, "page": page, "per_page": per_page}


@router.get("/audit")
async def audit_report(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[Session, Depends(get_db)],
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    module: Optional[str] = None,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    query = db.query(AuditLog)
    if module:
        query = query.filter(AuditLog.module == module)
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if action:
        query = query.filter(AuditLog.action == action)
    if date_from:
        try:
            query = query.filter(AuditLog.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(AuditLog.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            pass

    total = query.count()
    logs = (
        query.order_by(AuditLog.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    data = [
        {
            "id": log.id,
            "user_id": log.user_id,
            "user_name": log.user_name,
            "action": log.action,
            "module": log.module,
            "entity_id": log.entity_id,
            "entity_type": log.entity_type,
            "old_value": log.old_value,
            "new_value": log.new_value,
            "ip_address": log.ip_address,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]

    return {"data": data, "total": total, "page": page, "per_page": per_page}
